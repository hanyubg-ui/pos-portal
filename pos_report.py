#!/usr/bin/env python3
"""
pos_report.py — 小売POS売上レポート自動生成スクリプト

【入力】 CSV または Excel 形式の POS データ
【出力】 月×ブランド単位の Excel ファイル（商品ごとにシートを分割）

使い方:
    python pos_report.py <入力ファイル> [出力ディレクトリ] [オプション]

オプション:
    --split=brand    ブランドごとにファイルを作成し商品ごとにシートを追加（デフォルト・推奨）
    --split=product  商品ごとに別ファイルを出力
    --output=<dir>   出力先ディレクトリを指定（デフォルト: ./output）

サンプルデータ生成（テスト用）:
    python pos_report.py --sample [保存先パス.csv]
"""

import sys
import re
import calendar
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════════════
# スタイル設定
# ═══════════════════════════════════════════════════════════════════════

_S_THIN   = Side(style="thin",   color="AAAAAA")
_S_MEDIUM = Side(style="medium", color="444444")

BORDER_THIN = Border(
    left=_S_THIN, right=_S_THIN, top=_S_THIN, bottom=_S_THIN
)
BORDER_HEADER_BOTTOM = Border(
    left=_S_THIN, right=_S_THIN, top=_S_THIN, bottom=_S_MEDIUM
)

# カラーパレット
C_HEADER_DARK = "1F4E79"   # 濃い紺青 — メインヘッダー背景
C_HEADER_MID  = "2E75B6"   # 中間青   — 合計行ラベル背景
C_TOTAL_BG    = "DEEAF1"   # 薄い青   — 合計セル背景
C_RETAILER_BG = "EBF3FB"   # 極薄青   — 小売店名先頭行背景
C_ROW_ODD     = "F4F9FF"   # 交互縞（奇数行）
C_ROW_EVEN    = "FFFFFF"   # 交互縞（偶数行）
C_WHITE       = "FFFFFF"
C_BLACK       = "000000"
C_GRAY        = "888888"   # ゼロ値の文字色


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, size=10, color=C_BLACK, italic=False) -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic)


def _align(h="center", v="center", wrap=False, indent=0) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)


# ═══════════════════════════════════════════════════════════════════════
# データ読み込み・前処理
# ═══════════════════════════════════════════════════════════════════════

REQUIRED_COLS = {"日付", "小売店名", "店舗名", "ブランド名", "商品名", "売上数量"}

# PLAZAの生データ形式のブランド名マッピング（半角→全角変換後に前方一致）
_BRAND_PREFIXES = [
    ("ラフドット",   ["ラフドット"]),
    ("YOU TOKYO",    ["YOU TOKYO"]),
    ("TEABLESS",     ["TEABLESS"]),
    ("バイワンシー", ["バイワンシー"]),
    ("レチスパ",     ["レチスパ"]),
    ("Zフルーシー",  ["Zフルーシー", "Z フルーシー"]),
]


def _detect_brand(product_name: str) -> str:
    """商品名の先頭からブランド名を判定する"""
    for brand, prefixes in _BRAND_PREFIXES:
        for prefix in prefixes:
            if product_name.startswith(prefix):
                return brand
    return "（不明）"


def _is_loft_raw_format(df: pd.DataFrame) -> bool:
    """ロフトの生データ形式（ヘッダーなし12列・ゼロ埋め数量）かどうかを判定する"""
    if len(df.columns) < 9:
        return False
    try:
        # 列名が 0,1,2... の整数（ヘッダーなし）かつ先頭列がYYYYMMDD形式
        first_val = str(df.iloc[0, 0]).strip().strip('"')
        return first_val.isdigit() and len(first_val) == 8
    except Exception:
        return False


def _convert_loft_format(df: pd.DataFrame) -> pd.DataFrame:
    """ロフトの生データを標準縦形式に変換する"""
    import unicodedata

    df = df.copy()
    # 列インデックスで参照（ヘッダーなしのため）
    cols = df.columns.tolist()

    def clean(val):
        return str(val).strip().strip('"')

    # 日付: YYYYMMDD → YYYY-MM-DD
    dates = df.iloc[:, 0].apply(lambda x: clean(x)).apply(
        lambda x: f"{x[:4]}-{x[4:6]}-{x[6:8]}" if len(x) == 8 and x.isdigit() else None
    )

    # 店舗コード → 店舗名
    store_codes = df.iloc[:, 1].apply(clean)
    store_names = "ロフト" + store_codes

    # 商品名: NFKC正規化で半角→全角カナ変換
    product_names = df.iloc[:, 4].apply(
        lambda x: unicodedata.normalize("NFKC", clean(x))
    )

    # 数量・金額: ゼロ埋め文字列 → 整数
    quantities = df.iloc[:, 7].apply(lambda x: int(clean(x)) if clean(x).isdigit() else 0)
    amounts    = df.iloc[:, 8].apply(lambda x: int(clean(x)) if clean(x).isdigit() else 0)

    # ブランド名を商品名から判定
    brands = product_names.apply(_detect_brand)

    result = pd.DataFrame({
        "日付":     dates,
        "小売店名": "ロフト",
        "店舗名":   store_names,
        "ブランド名": brands,
        "商品名":   product_names,
        "売上数量": quantities,
        "売上金額": amounts,
    })

    # 数量0・日付なしを除外
    result = result[result["売上数量"] > 0].dropna(subset=["日付"]).reset_index(drop=True)
    return result


def _is_plaza_raw_format(df: pd.DataFrame) -> bool:
    """PLAZAの生データ形式（日別横並び）かどうかを判定する"""
    return "データ更新年月日" in df.columns and "２８日間売数" in df.columns


def _convert_plaza_format(df: pd.DataFrame) -> pd.DataFrame:
    """PLAZAの生データを標準縦形式に変換する。
    Excel形式（X月Y日列）とCSV形式（売上数N日前列）の両方に対応。"""
    import unicodedata

    df = df.copy()

    report_date_str = str(df["データ更新年月日"].dropna().iloc[0]).strip()[:8]
    report_date = pd.to_datetime(report_date_str, format="%Y%m%d")
    year = report_date.year

    def _norm(s):
        return unicodedata.normalize("NFKC", str(s).strip())

    # パターンA: Excel形式 "4月1日"
    re_a = re.compile(r'^(\d+)月(\d+)日$')
    day_cols_a = [c for c in df.columns if re_a.match(str(c).strip())]

    # パターンB: CSV形式 "売上数１日前"（全角数字を正規化して判定）
    re_b = re.compile(r'^売上数(\d+)日前$')
    day_cols_b = [c for c in df.columns if re_b.match(_norm(c))]

    if day_cols_a:
        day_cols = day_cols_a
        qty28 = pd.to_numeric(df["２８日間売数"], errors="coerce").fillna(0)
        amt28 = pd.to_numeric(df["２８日間売上金額"], errors="coerce").fillna(0)
        df["_unit_price"] = 0
        mask = qty28 > 0
        df.loc[mask, "_unit_price"] = (amt28[mask] / qty28[mask]).round(0).astype(int)

        melted = df[["店舗名", "商品名", "_unit_price"] + day_cols].melt(
            id_vars=["店舗名", "商品名", "_unit_price"],
            value_vars=day_cols, var_name="_date_str", value_name="売上数量",
        )
        def _parse_a(s):
            m = re_a.match(str(s).strip())
            return f"{year}-{int(m.group(1)):02d}-{int(m.group(2)):02d}" if m else None
        melted["日付"] = melted["_date_str"].apply(_parse_a)
        melted["売上金額"] = (melted["売上数量"].apply(
            lambda x: pd.to_numeric(x, errors="coerce") or 0) * melted["_unit_price"]).astype(int)

    elif day_cols_b:
        day_cols = day_cols_b
        if "上代単価" in df.columns:
            df["_unit_price"] = pd.to_numeric(df["上代単価"], errors="coerce").fillna(0).astype(int)
        else:
            qty28 = pd.to_numeric(df["２８日間売数"], errors="coerce").fillna(0)
            amt28 = pd.to_numeric(df["２８日間売上金額"], errors="coerce").fillna(0)
            df["_unit_price"] = 0
            mask = qty28 > 0
            df.loc[mask, "_unit_price"] = (amt28[mask] / qty28[mask]).round(0).astype(int)

        melted = df[["店舗名", "商品名", "_unit_price"] + day_cols].melt(
            id_vars=["店舗名", "商品名", "_unit_price"],
            value_vars=day_cols, var_name="_date_str", value_name="売上数量",
        )
        def _parse_b(s):
            m = re_b.match(_norm(s))
            if m:
                n = int(m.group(1))
                return (report_date - pd.Timedelta(days=n - 1)).strftime("%Y-%m-%d")
            return None
        melted["日付"] = melted["_date_str"].apply(_parse_b)
        melted["売上金額"] = (
            pd.to_numeric(melted["売上数量"], errors="coerce").fillna(0) * melted["_unit_price"]
        ).astype(int)

    else:
        raise ValueError("日別列（'4月1日' または '売上数N日前' 形式）が見つかりません。")

    melted["売上数量"] = pd.to_numeric(melted["売上数量"], errors="coerce").fillna(0).astype(int)
    melted = melted[melted["売上数量"] > 0].dropna(subset=["日付"]).copy()
    melted["商品名"]     = melted["商品名"].apply(lambda x: _norm(x))
    melted["ブランド名"] = melted["商品名"].apply(_detect_brand)
    melted["小売店名"]   = "PLAZA"
    melted["店舗名"]     = melted["店舗名"].apply(lambda x: _norm(x))

    return melted[["日付", "小売店名", "店舗名", "ブランド名", "商品名", "売上数量", "売上金額"]].reset_index(drop=True)


def load_pos_data(filepath: str | Path) -> pd.DataFrame:
    """POSデータを読み込み、型変換・バリデーションを行う。
    PLAZAの生データ形式（日別横並び）は自動的に標準形式に変換する。"""
    p = Path(filepath)

    if p.suffix.lower() in (".xlsx", ".xls"):
        df = pd.read_excel(p, dtype=str)
    elif p.suffix.lower() == ".csv":
        for enc in ("cp932", "utf-8-sig", "utf-8"):
            try:
                df = pd.read_csv(p, dtype=str, encoding=enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError(f"文字コードの自動判定に失敗しました: {p}")
    else:
        raise ValueError(f"未対応のファイル形式: {p.suffix}（.xlsx / .xls / .csv のみ対応）")

    # カラム名の前後の空白を除去
    df.columns = df.columns.str.strip()

    # ロフト生データ形式を自動検出・変換
    if _is_loft_raw_format(df):
        df = _convert_loft_format(df)
    # PLAZAの生データ形式を自動検出・変換
    elif _is_plaza_raw_format(df):
        df = _convert_plaza_format(df)

    # 必須カラムの確認
    missing = REQUIRED_COLS - set(df.columns)
    if missing:
        raise ValueError(f"必須カラムが見つかりません: {sorted(missing)}\n実際のカラム: {list(df.columns)}")

    # 型変換
    df["日付"]    = pd.to_datetime(df["日付"], errors="coerce")
    df            = df.dropna(subset=["日付"])
    df["売上数量"] = pd.to_numeric(df["売上数量"], errors="coerce").fillna(0).astype(int)

    if "売上金額" not in df.columns:
        df["売上金額"] = 0
    df["売上金額"] = pd.to_numeric(df["売上金額"], errors="coerce").fillna(0).astype(int)

    for col in ("小売店名", "店舗名", "ブランド名", "商品名"):
        df[col] = df[col].astype(str).str.strip().replace("nan", "（不明）").fillna("（不明）")

    return df


def _sanitize_sheet_name(name: str) -> str:
    """Excelシート名に使えない文字を除去（最大31文字）"""
    cleaned = re.sub(r'[\\/:*?\[\]]', "_", str(name))
    return cleaned[:31]


def _sanitize_filename(name: str) -> str:
    """ファイル名に使えない文字を除去"""
    return re.sub(r'[\\/:*?"<>|]', "_", str(name))


# ═══════════════════════════════════════════════════════════════════════
# Excelシート書き込み
# ═══════════════════════════════════════════════════════════════════════

def write_product_sheet(
    ws,
    df_prod: pd.DataFrame,
    product_name: str,
    brand_name: str,
    year: int,
    month: int,
) -> None:
    """
    1商品分のクロス集計シートを書き込む

    レイアウト（行方向）:
        row 1   : タイトル行（商品名・ブランド名・年月）
        row 2   : カラムヘッダー（小売店名 | 店舗名 | 1〜末日 | 合計）
        row 3〜 : データ行（小売店が変わる行のみ小売店名を表示）
        最終行   : 日計合計行
    """
    last_day = calendar.monthrange(year, month)[1]
    all_days = list(range(1, last_day + 1))
    n_days   = len(all_days)

    # ── ピボットテーブル生成 ─────────────────────────────────────
    df_prod       = df_prod.copy()
    df_prod["日"] = df_prod["日付"].dt.day

    pivot = df_prod.pivot_table(
        index=["小売店名", "店舗名"],
        columns="日",
        values="売上数量",
        aggfunc="sum",
        fill_value=0,
    )
    # 月の全日（データがない日も 0 で補完）
    pivot = pivot.reindex(columns=all_days, fill_value=0)

    # 行合計（月間累計）を追加
    pivot["合計"] = pivot[all_days].sum(axis=1)

    # 列合計（日計）
    day_totals  = pivot[all_days].sum(axis=0)
    grand_total = int(pivot["合計"].sum())

    # ── 列番号の定義 ─────────────────────────────────────────────
    COL_RETAILER  = 1                        # 小売店名
    COL_STORE     = 2                        # 店舗名
    COL_DAY_START = 3                        # 日付列の開始
    COL_TOTAL     = COL_DAY_START + n_days   # 合計列

    ROW_TITLE  = 1
    ROW_HEADER = 2
    ROW_DATA   = 3

    # ── row 1: タイトル ──────────────────────────────────────────
    title_text = f"{year}年{month}月  {brand_name}  {product_name}  売上数量集計"
    tc = ws.cell(ROW_TITLE, COL_RETAILER, title_text)
    tc.font      = _font(bold=True, size=12, color=C_HEADER_DARK)
    tc.alignment = _align(h="left", v="center")
    ws.merge_cells(
        start_row=ROW_TITLE, start_column=COL_RETAILER,
        end_row=ROW_TITLE,   end_column=COL_TOTAL,
    )
    ws.row_dimensions[ROW_TITLE].height = 22

    # ── row 2: カラムヘッダー ─────────────────────────────────────
    for col, label in [(COL_RETAILER, "小売店名"), (COL_STORE, "店舗名")]:
        c = ws.cell(ROW_HEADER, col, label)
        c.font      = _font(bold=True, color=C_WHITE)
        c.fill      = _fill(C_HEADER_DARK)
        c.alignment = _align(h="center")
        c.border    = BORDER_HEADER_BOTTOM

    for i, day in enumerate(all_days):
        c = ws.cell(ROW_HEADER, COL_DAY_START + i, day)
        c.font      = _font(bold=True, color=C_WHITE, size=9)
        c.fill      = _fill(C_HEADER_DARK)
        c.alignment = _align(h="center")
        c.border    = BORDER_HEADER_BOTTOM

    c = ws.cell(ROW_HEADER, COL_TOTAL, "合計")
    c.font      = _font(bold=True, color=C_WHITE)
    c.fill      = _fill(C_HEADER_MID)
    c.alignment = _align(h="center")
    c.border    = BORDER_HEADER_BOTTOM

    ws.row_dimensions[ROW_HEADER].height = 18

    # ── データ行 ─────────────────────────────────────────────────
    prev_retailer = None
    row_idx = ROW_DATA

    for (retailer, store), row_data in pivot.iterrows():
        is_odd  = ((row_idx - ROW_DATA) % 2 == 0)
        row_bg  = C_ROW_ODD if is_odd else C_ROW_EVEN
        is_new_retailer = (retailer != prev_retailer)

        # 小売店名セル（グループ先頭のみ値を表示）
        c = ws.cell(ROW_HEADER, COL_RETAILER)  # ダミー参照
        c = ws.cell(row_idx, COL_RETAILER, retailer if is_new_retailer else "")
        c.font      = _font(bold=is_new_retailer, size=10)
        c.fill      = _fill(C_RETAILER_BG if is_new_retailer else C_ROW_EVEN)
        c.alignment = _align(h="left", v="center")
        c.border    = BORDER_THIN
        prev_retailer = retailer

        # 店舗名セル
        c = ws.cell(row_idx, COL_STORE, store)
        c.font      = _font(size=10)
        c.fill      = _fill(row_bg)
        c.alignment = _align(h="left", indent=1)
        c.border    = BORDER_THIN

        # 日付データセル
        for i, day in enumerate(all_days):
            val = int(row_data[day])
            c = ws.cell(row_idx, COL_DAY_START + i, val)
            c.font         = _font(size=10, color=C_GRAY if val == 0 else C_BLACK)
            c.fill         = _fill(row_bg)
            c.alignment    = _align(h="right")
            c.border       = BORDER_THIN
            c.number_format = "#,##0"

        # 行合計（月間累計）セル
        row_total = int(row_data["合計"])
        c = ws.cell(row_idx, COL_TOTAL, row_total)
        c.font         = _font(bold=True, size=10)
        c.fill         = _fill(C_TOTAL_BG)
        c.alignment    = _align(h="right")
        c.border       = BORDER_THIN
        c.number_format = "#,##0"

        ws.row_dimensions[row_idx].height = 16
        row_idx += 1

    # ── 最終行: 日計合計 ──────────────────────────────────────────
    ROW_TOTAL = row_idx

    c = ws.cell(ROW_TOTAL, COL_RETAILER, "合計")
    c.font      = _font(bold=True, color=C_WHITE)
    c.fill      = _fill(C_HEADER_MID)
    c.alignment = _align(h="center")
    c.border    = BORDER_THIN

    c = ws.cell(ROW_TOTAL, COL_STORE, "")
    c.fill   = _fill(C_HEADER_MID)
    c.border = BORDER_THIN

    for i, day in enumerate(all_days):
        val = int(day_totals.get(day, 0))
        c = ws.cell(ROW_TOTAL, COL_DAY_START + i, val)
        c.font         = _font(bold=True, size=10)
        c.fill         = _fill(C_TOTAL_BG)
        c.alignment    = _align(h="right")
        c.border       = BORDER_THIN
        c.number_format = "#,##0"

    # 右下の総合計セル（目立つスタイル）
    c = ws.cell(ROW_TOTAL, COL_TOTAL, grand_total)
    c.font         = _font(bold=True, color=C_WHITE, size=11)
    c.fill         = _fill(C_HEADER_DARK)
    c.alignment    = _align(h="right")
    c.border       = BORDER_THIN
    c.number_format = "#,##0"

    ws.row_dimensions[ROW_TOTAL].height = 18

    # ── 列幅の自動調整 ────────────────────────────────────────────
    # 日本語全角文字は約2文字分の幅を持つため係数 1.8 を掛けて近似
    max_retailer_len = max(
        (len(str(v)) for v in pivot.index.get_level_values(0)), default=6
    )
    max_store_len = max(
        (len(str(v)) for v in pivot.index.get_level_values(1)), default=6
    )
    ws.column_dimensions[get_column_letter(COL_RETAILER)].width = max(max_retailer_len * 1.8 + 2, 14)
    ws.column_dimensions[get_column_letter(COL_STORE)].width    = max(max_store_len    * 1.8 + 2, 14)

    for i in range(n_days):
        ws.column_dimensions[get_column_letter(COL_DAY_START + i)].width = 5

    ws.column_dimensions[get_column_letter(COL_TOTAL)].width = 9

    # ── ウィンドウ枠の固定（ヘッダー2行＋左2列）────────────────────
    ws.freeze_panes = ws.cell(ROW_DATA, COL_DAY_START)

    # シート名（31文字制限）
    ws.title = _sanitize_sheet_name(product_name)


# ═══════════════════════════════════════════════════════════════════════
# レポート生成メイン処理
# ═══════════════════════════════════════════════════════════════════════

def generate_reports(
    df: pd.DataFrame,
    output_dir: Path,
    split_mode: str = "brand",
) -> list[Path]:
    """
    月×ブランド単位でExcelレポートを生成する

    split_mode:
        "brand"   : ブランドごとにファイルを作成、商品ごとにシートを追加（デフォルト）
        "product" : 商品ごとに別ファイルを出力
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    df = df.copy()
    df["_period"] = df["日付"].dt.to_period("M")

    generated: list[Path] = []

    for (period, brand), df_brand in df.groupby(["_period", "ブランド名"], sort=True):
        year  = period.year
        month = period.month
        ym    = f"{year}年{month}月"
        brand_safe = _sanitize_filename(brand)
        products   = sorted(df_brand["商品名"].unique())

        if split_mode == "product":
            # ── 商品ごとに別ファイル ──────────────────────────────
            for product in products:
                df_prod   = df_brand[df_brand["商品名"] == product]
                prod_safe = _sanitize_filename(product)
                fpath     = output_dir / f"{ym}_{brand_safe}_{prod_safe}_売上.xlsx"

                wb = Workbook()
                ws = wb.active
                write_product_sheet(ws, df_prod, product, brand, year, month)

                wb.save(fpath)
                generated.append(fpath)
                print(f"  保存: {fpath.name}")

        else:
            # ── ブランドごとにファイル、商品ごとにシート（デフォルト）──
            fpath = output_dir / f"{ym}_{brand_safe}_売上.xlsx"

            wb = Workbook()
            wb.remove(wb.active)  # デフォルト空シートを削除

            for product in products:
                df_prod = df_brand[df_brand["商品名"] == product]
                ws = wb.create_sheet()
                write_product_sheet(ws, df_prod, product, brand, year, month)

            wb.save(fpath)
            generated.append(fpath)
            sheet_names = [s.title for s in wb.worksheets]
            print(f"  保存: {fpath.name}  [{len(products)} シート: {', '.join(sheet_names)}]")

    return generated


# ═══════════════════════════════════════════════════════════════════════
# サンプルデータ生成（テスト・動作確認用）
# ═══════════════════════════════════════════════════════════════════════

def create_sample_data(filepath: str | Path) -> None:
    """動作確認用サンプルPOSデータを生成する"""
    import random
    import numpy as np

    random.seed(42)
    np.random.seed(42)

    retailers = {
        "PLAZA":  ["渋谷店", "新宿店", "原宿店"],
        "ハンズ": ["渋谷店", "梅田店"],
        "ロフト": ["銀座店", "池袋店", "名古屋店"],
    }
    brands = {
        "ブランドA": ["商品X", "商品Y", "商品Z"],
        "ブランドB": ["商品P", "商品Q"],
    }
    months = [
        ("2026-04", 30),
        ("2026-05", 31),
    ]

    rows = []
    for ym, days in months:
        for day in range(1, days + 1):
            date = f"{ym}-{day:02d}"
            for brand, products in brands.items():
                for product in products:
                    for retailer, stores in retailers.items():
                        # 各店舗をランダムにサンプリング（売上がない日もある）
                        for store in random.sample(stores, k=random.randint(1, len(stores))):
                            qty = random.choices(
                                [0, 1, 2, 3, 5, 8],
                                weights=[35, 25, 18, 12, 7, 3],
                            )[0]
                            if qty > 0:
                                rows.append({
                                    "日付":       date,
                                    "小売店名":   retailer,
                                    "店舗名":     store,
                                    "ブランド名": brand,
                                    "商品名":     product,
                                    "売上数量":   qty,
                                    "売上金額":   qty * random.randint(800, 3000),
                                })

    df = pd.DataFrame(rows)
    p  = Path(filepath)
    if p.suffix.lower() == ".csv":
        df.to_csv(p, index=False, encoding="utf-8-sig")
    else:
        df.to_excel(p, index=False)

    print(f"サンプルデータを生成しました: {p}  ({len(df):,} 件)")
    print(f"  期間: {df['日付'].min()} 〜 {df['日付'].max()}")
    print(f"  ブランド: {sorted(df['ブランド名'].unique())}")
    print(f"  商品数: {df['商品名'].nunique()}")
    print(f"  小売店: {sorted(df['小売店名'].unique())}")


# ═══════════════════════════════════════════════════════════════════════
# CLIエントリーポイント
# ═══════════════════════════════════════════════════════════════════════

def main():
    args = sys.argv[1:]

    if not args or args[0] in ("-h", "--help"):
        print(__doc__)
        sys.exit(0)

    # ─── --sample モード ─────────────────────────────────────────
    if args[0] == "--sample":
        sample_path = args[1] if len(args) > 1 else "sample_pos.csv"
        create_sample_data(sample_path)
        return

    # ─── 通常モード ──────────────────────────────────────────────
    input_file = Path(args[0])
    output_dir = Path("output")
    split_mode = "brand"

    for arg in args[1:]:
        if arg.startswith("--output="):
            output_dir = Path(arg.split("=", 1)[1])
        elif arg == "--split=product":
            split_mode = "product"
        elif arg == "--split=brand":
            split_mode = "brand"
        elif not arg.startswith("--"):
            output_dir = Path(arg)

    print(f"入力ファイル : {input_file}")
    df = load_pos_data(input_file)

    print(f"  読み込み  : {len(df):,} 件")
    print(f"  期間      : {df['日付'].min().date()} 〜 {df['日付'].max().date()}")
    print(f"  ブランド  : {sorted(df['ブランド名'].unique())}")
    print(f"  商品数    : {df['商品名'].nunique()}")

    mode_label = "商品ごとにファイル分割" if split_mode == "product" else "ブランドごとにファイル分割（商品ごとシート）"
    print(f"\n出力先      : {output_dir}")
    print(f"分割モード  : {mode_label}")
    print()

    files = generate_reports(df, output_dir, split_mode=split_mode)

    print(f"\n完了: {len(files)} ファイルを {output_dir.resolve()} に生成しました。")


if __name__ == "__main__":
    main()
